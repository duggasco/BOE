"""
Celery tasks for report exports (CSV, Excel, PDF).
"""
import os
import csv
import json
import tempfile
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
from pathlib import Path
import uuid

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from celery import Task
from celery.utils.log import get_task_logger
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.celery_app import celery_app
from app.core.database import async_session_maker
from app.models.report import Report, ReportExecution
from app.models.user import User
from app.services.query_builder import QueryBuilder
from app.services.report_service import ReportService

logger = get_task_logger(__name__)


class ExportTask(Task):
    """Base class for export tasks with common functionality."""
    
    def on_failure(self, exc, task_id, args, kwargs, einfo):
        """Handle task failure."""
        logger.error(f"Export task {task_id} failed: {exc}")
        # Update execution status in database
        execution_id = kwargs.get("execution_id")
        if execution_id:
            # This would need to be implemented with proper async handling
            pass
    
    def on_success(self, retval, task_id, args, kwargs):
        """Handle task success."""
        logger.info(f"Export task {task_id} completed successfully")


@celery_app.task(base=ExportTask, bind=True, name="export.csv")
def export_to_csv(
    self,
    report_id: str,
    user_id: str,
    execution_id: str,
    parameters: Optional[Dict[str, Any]] = None,
    options: Optional[Dict[str, Any]] = None
) -> Dict[str, Any]:
    """
    Export report data to CSV format.
    
    Args:
        report_id: ID of the report to export
        user_id: ID of the user requesting export
        execution_id: ID of the execution record
        parameters: Report parameters for execution
        options: Export options (delimiter, encoding, etc.)
    
    Returns:
        Dict with export status and file path
    """
    try:
        logger.info(f"Starting CSV export for report {report_id}")
        
        # Default options
        options = options or {}
        delimiter = options.get("delimiter", ",")
        encoding = options.get("encoding", "utf-8")
        include_headers = options.get("include_headers", True)
        
        # Create temporary file
        temp_dir = Path(tempfile.gettempdir()) / "boe_exports"
        temp_dir.mkdir(exist_ok=True)
        
        filename = f"report_{report_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = temp_dir / filename
        
        # Execute report query synchronously (simplified for this implementation)
        # In production, this would use async database operations
        data = execute_report_query_sync(report_id, parameters)
        
        # Write CSV file
        with open(filepath, 'w', newline='', encoding=encoding) as csvfile:
            if not data:
                # Empty result
                csvfile.write("No data available\n")
            else:
                # Get column names from first row
                columns = list(data[0].keys()) if data else []
                
                writer = csv.DictWriter(
                    csvfile,
                    fieldnames=columns,
                    delimiter=delimiter
                )
                
                if include_headers:
                    writer.writeheader()
                
                writer.writerows(data)
        
        # Update task progress
        self.update_state(
            state='SUCCESS',
            meta={
                'current': 100,
                'total': 100,
                'status': 'Export completed',
                'filepath': str(filepath)
            }
        )
        
        logger.info(f"CSV export completed: {filepath}")
        
        return {
            "status": "success",
            "filepath": str(filepath),
            "filename": filename,
            "size": os.path.getsize(filepath),
            "rows": len(data)
        }
        
    except Exception as e:
        logger.error(f"CSV export failed: {str(e)}")
        raise


@celery_app.task(base=ExportTask, bind=True, name="export.excel")
def export_to_excel(
    self,
    report_id: str,
    user_id: str,
    execution_id: str,
    parameters: Optional[Dict[str, Any]] = None,
    options: Optional[Dict[str, Any]] = None
) -> Dict[str, Any]:
    """
    Export report data to Excel format.
    
    Args:
        report_id: ID of the report to export
        user_id: ID of the user requesting export
        execution_id: ID of the execution record
        parameters: Report parameters for execution
        options: Export options (formatting, multiple sheets, etc.)
    
    Returns:
        Dict with export status and file path
    """
    try:
        logger.info(f"Starting Excel export for report {report_id}")
        
        # Default options
        options = options or {}
        include_formatting = options.get("include_formatting", True)
        freeze_headers = options.get("freeze_headers", True)
        auto_filter = options.get("auto_filter", True)
        
        # Create temporary file
        temp_dir = Path(tempfile.gettempdir()) / "boe_exports"
        temp_dir.mkdir(exist_ok=True)
        
        filename = f"report_{report_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = temp_dir / filename
        
        # Execute report query
        data = execute_report_query_sync(report_id, parameters)
        
        # Create Excel workbook
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Report Data"
        
        if not data:
            worksheet.append(["No data available"])
        else:
            # Convert to DataFrame for easier handling
            df = pd.DataFrame(data)
            
            # Add headers with formatting
            headers = list(df.columns)
            worksheet.append(headers)
            
            if include_formatting:
                # Style headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(
                    start_color="366092",
                    end_color="366092",
                    fill_type="solid"
                )
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for col_num, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
            
            # Add data rows
            for row in dataframe_to_rows(df, index=False, header=False):
                worksheet.append(row)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze top row
            if freeze_headers:
                worksheet.freeze_panes = "A2"
            
            # Add auto filter
            if auto_filter:
                worksheet.auto_filter.ref = worksheet.dimensions
        
        # Add metadata sheet
        metadata_sheet = workbook.create_sheet("Metadata")
        metadata_sheet.append(["Report ID", report_id])
        metadata_sheet.append(["Generated At", datetime.utcnow().isoformat()])
        metadata_sheet.append(["Generated By", user_id])
        metadata_sheet.append(["Parameters", json.dumps(parameters or {})])
        metadata_sheet.append(["Total Rows", len(data)])
        
        # Save workbook
        workbook.save(filepath)
        
        # Update task progress
        self.update_state(
            state='SUCCESS',
            meta={
                'current': 100,
                'total': 100,
                'status': 'Export completed',
                'filepath': str(filepath)
            }
        )
        
        logger.info(f"Excel export completed: {filepath}")
        
        return {
            "status": "success",
            "filepath": str(filepath),
            "filename": filename,
            "size": os.path.getsize(filepath),
            "rows": len(data)
        }
        
    except Exception as e:
        logger.error(f"Excel export failed: {str(e)}")
        raise


@celery_app.task(base=ExportTask, bind=True, name="export.pdf")
def export_to_pdf(
    self,
    report_id: str,
    user_id: str,
    execution_id: str,
    parameters: Optional[Dict[str, Any]] = None,
    options: Optional[Dict[str, Any]] = None
) -> Dict[str, Any]:
    """
    Export report data to PDF format.
    
    Args:
        report_id: ID of the report to export
        user_id: ID of the user requesting export
        execution_id: ID of the execution record
        parameters: Report parameters for execution
        options: Export options (page size, orientation, etc.)
    
    Returns:
        Dict with export status and file path
    """
    try:
        logger.info(f"Starting PDF export for report {report_id}")
        
        # Default options
        options = options or {}
        page_size = options.get("page_size", "letter")
        orientation = options.get("orientation", "portrait")
        include_header = options.get("include_header", True)
        include_footer = options.get("include_footer", True)
        
        # Create temporary file
        temp_dir = Path(tempfile.gettempdir()) / "boe_exports"
        temp_dir.mkdir(exist_ok=True)
        
        filename = f"report_{report_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = temp_dir / filename
        
        # Execute report query
        data = execute_report_query_sync(report_id, parameters)
        report_info = get_report_info_sync(report_id)
        
        # Create PDF document
        if page_size == "letter":
            pagesize = letter
        else:
            pagesize = A4
        
        doc = SimpleDocTemplate(
            str(filepath),
            pagesize=pagesize,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18,
        )
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Add title
        if include_header:
            title = Paragraph(report_info.get("name", "Report"), title_style)
            elements.append(title)
            
            # Add report description
            if report_info.get("description"):
                desc = Paragraph(report_info["description"], styles['Normal'])
                elements.append(desc)
                elements.append(Spacer(1, 12))
            
            # Add generation info
            gen_info = Paragraph(
                f"Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}",
                styles['Normal']
            )
            elements.append(gen_info)
            elements.append(Spacer(1, 20))
        
        # Create table from data
        if not data:
            no_data = Paragraph("No data available", styles['Normal'])
            elements.append(no_data)
        else:
            # Prepare table data
            df = pd.DataFrame(data)
            
            # Add headers
            table_data = [list(df.columns)]
            
            # Add data rows (limit rows for PDF to prevent huge files)
            max_rows = options.get("max_rows", 1000)
            for _, row in df.head(max_rows).iterrows():
                table_data.append([str(val) for val in row.values])
            
            # Create table
            table = Table(table_data)
            
            # Apply table style
            table_style = TableStyle([
                # Header style
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Data style
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ])
            
            table.setStyle(table_style)
            elements.append(table)
            
            # Add note if data was truncated
            if len(df) > max_rows:
                truncation_note = Paragraph(
                    f"Note: Data truncated to {max_rows} rows for PDF export. "
                    f"Total rows: {len(df)}",
                    styles['Italic']
                )
                elements.append(Spacer(1, 12))
                elements.append(truncation_note)
        
        # Add footer
        if include_footer:
            elements.append(Spacer(1, 20))
            footer = Paragraph(
                f"Page <pageNumber/> | Report ID: {report_id}",
                styles['Normal']
            )
            elements.append(footer)
        
        # Build PDF
        doc.build(elements)
        
        # Update task progress
        self.update_state(
            state='SUCCESS',
            meta={
                'current': 100,
                'total': 100,
                'status': 'Export completed',
                'filepath': str(filepath)
            }
        )
        
        logger.info(f"PDF export completed: {filepath}")
        
        return {
            "status": "success",
            "filepath": str(filepath),
            "filename": filename,
            "size": os.path.getsize(filepath),
            "rows": len(data) if data else 0
        }
        
    except Exception as e:
        logger.error(f"PDF export failed: {str(e)}")
        raise


@celery_app.task(name="export.cleanup")
def cleanup_old_exports():
    """
    Cleanup old export files to prevent disk space issues.
    Runs periodically via Celery beat.
    """
    try:
        logger.info("Starting cleanup of old export files")
        
        temp_dir = Path(tempfile.gettempdir()) / "boe_exports"
        if not temp_dir.exists():
            return {"status": "success", "message": "No export directory found"}
        
        # Remove files older than 24 hours
        cutoff_time = datetime.now() - timedelta(hours=24)
        removed_count = 0
        
        for filepath in temp_dir.glob("report_*"):
            if filepath.is_file():
                file_time = datetime.fromtimestamp(filepath.stat().st_mtime)
                if file_time < cutoff_time:
                    filepath.unlink()
                    removed_count += 1
                    logger.info(f"Removed old export file: {filepath}")
        
        logger.info(f"Cleanup completed. Removed {removed_count} files")
        
        return {
            "status": "success",
            "removed_count": removed_count
        }
        
    except Exception as e:
        logger.error(f"Cleanup failed: {str(e)}")
        raise


# Helper functions (simplified for this implementation)
def execute_report_query_sync(report_id: str, parameters: Dict[str, Any] = None) -> List[Dict[str, Any]]:
    """
    Execute report query synchronously.
    In production, this would use proper async database operations.
    """
    # Simplified mock implementation
    # In reality, this would:
    # 1. Fetch report definition from database
    # 2. Build SQL query using QueryBuilder
    # 3. Execute query with parameters
    # 4. Return results as list of dicts
    
    # Mock data for testing
    return [
        {
            "id": i,
            "name": f"Item {i}",
            "value": i * 100,
            "date": datetime.utcnow().isoformat(),
            "category": f"Category {i % 5}"
        }
        for i in range(1, 101)
    ]


def get_report_info_sync(report_id: str) -> Dict[str, Any]:
    """
    Get report information synchronously.
    In production, this would fetch from database.
    """
    # Mock implementation
    return {
        "id": report_id,
        "name": "Sample Report",
        "description": "This is a sample report for export testing",
        "created_at": datetime.utcnow().isoformat()
    }